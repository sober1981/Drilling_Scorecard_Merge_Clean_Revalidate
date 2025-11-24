#!/usr/bin/env python3
"""
REVALIDATE MASTER Script
Version: 1.0
Date: 2025-11-24

This script re-validates a MERGE_CLEAN_QC_MASTER file after it's been updated with new runs.
It applies the same QC logic as 5_qc_revalidate.py but is specifically designed for MASTER files.

IMPORTANT FEATURES:
1. Respects manual changes (if you removed yellow highlighting or edited cells)
2. Preserves the 4 special columns (REASON_POOH_QC, SERIES 20, CONTROL #, QC BY)
3. Does NOT validate or highlight the 4 special columns
4. Creates output: MERGE_CLEAN_QC_MASTER_REVALIDATED_[timestamp].xlsx

The MASTER_REVALIDATED file becomes the new MASTER for the next merge cycle.

Usage:
1. Run Script 6 to create MERGE_CLEAN_QC_MASTER_[timestamp].xlsx
2. Optionally make manual corrections to the file
3. Run this script to create MERGE_CLEAN_QC_MASTER_REVALIDATED_[timestamp].xlsx
4. Use the REVALIDATED file as the MASTER for the next merge
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import os
import glob

# Import validation logic from 4_qc_data_quality
import importlib.util
import sys

# Load the QC module
spec = importlib.util.spec_from_file_location("qc_data_quality", "4_qc_data_quality.py")
qc_module = importlib.util.module_from_spec(spec)
sys.modules["qc_data_quality"] = qc_module
spec.loader.exec_module(qc_module)

# Import the functions we need
load_qc_criteria = qc_module.load_qc_criteria
check_cell = qc_module.check_cell

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
NO_FILL = PatternFill(fill_type=None)

# Special columns that should NOT be validated or highlighted
SPECIAL_COLUMNS = ['REASON_POOH_QC', 'SERIES 20', 'CONTROL #', 'QC BY']


def find_master_file():
    """
    Find the most recent MERGE_CLEAN_QC_MASTER file in the current directory.

    IMPORTANT: This includes REVALIDATED files because the workflow is:
    1. Script 6 creates MERGE_CLEAN_QC_MASTER_[timestamp].xlsx
    2. Script 7 creates MERGE_CLEAN_QC_MASTER_REVALIDATED_[timestamp].xlsx
    3. User edits REVALIDATED file manually
    4. Script 7 should read the REVALIDATED file (the most recent MASTER)
    """
    print("\nSearching for MERGE_CLEAN_QC_MASTER file...")
    pattern = "MERGE_CLEAN_QC_MASTER*.xlsx"
    matches = glob.glob(pattern)

    # DO NOT exclude REVALIDATED files - they are the MASTER for next iteration!
    # Sort by modification time, get most recent
    if len(matches) == 0:
        print(f"  ERROR: No file found matching pattern '{pattern}'")
        return None
    elif len(matches) > 1:
        matches.sort(key=os.path.getmtime, reverse=True)
        print(f"  Multiple files found. Using most recent: {matches[0]}")
    else:
        print(f"  Found: {matches[0]}")

    return matches[0]


def get_existing_highlights(input_file):
    """
    Read the input file and identify which cells currently have yellow highlighting.
    Returns: set of (row_idx, col_name) tuples for cells WITHOUT yellow highlighting

    IMPORTANT: Cells without yellow are considered "manually QC'd" and won't be re-flagged.
    """
    print("\nReading existing cell highlighting from file...")

    wb = load_workbook(input_file)
    ws = wb.active

    # Get column names from header row
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_names = list(header_row)

    # Track cells that DON'T have yellow (these are either fixed or manually QC'd)
    non_yellow_cells = set()
    cells_with_yellow = 0
    cells_without_yellow = 0

    # Iterate through all data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row)):
        for col_idx, cell in enumerate(row):
            col_name = col_names[col_idx]

            # Skip special columns - never consider them for highlighting
            if col_name in SPECIAL_COLUMNS:
                non_yellow_cells.add((row_idx, col_name))
                continue

            # Check if cell has yellow highlighting
            has_yellow = False

            if cell.fill and cell.fill.start_color:
                fill_color = str(cell.fill.start_color.rgb) if hasattr(cell.fill.start_color, 'rgb') else str(cell.fill.start_color.index)

                # Check if it's yellow (FFFF00 or variations)
                is_yellow = (
                    'FFFF00' in fill_color.upper() or
                    'FFFFFF00' in fill_color.upper() or
                    fill_color == 'FFFF0000'  # Some Excel versions use this
                )

                if is_yellow:
                    has_yellow = True
                    cells_with_yellow += 1

            if not has_yellow:
                # Cell doesn't have yellow - consider it manually QC'd
                non_yellow_cells.add((row_idx, col_name))
                cells_without_yellow += 1

    print(f"  Found {cells_with_yellow} cells WITH yellow highlighting")
    print(f"  Found {cells_without_yellow} cells WITHOUT yellow (considered manually QC'd)")
    print(f"  Note: {len(SPECIAL_COLUMNS)} special columns are excluded from QC validation")

    # Debug: Show some examples of non-yellow cells
    if len(non_yellow_cells) > 0:
        sample_count = min(5, len(non_yellow_cells))
        sample_cells = list(non_yellow_cells)[:sample_count]
        print(f"\n  Sample of non-yellow cells (first {sample_count}):")
        for row_idx, col_name in sample_cells:
            print(f"    Row {row_idx + 1}, Column '{col_name}'")

    wb.close()
    return non_yellow_cells


def validate_data(df, criteria_dict, phase_map):
    """
    Validate entire dataframe against QC criteria.
    Returns: dict of issues {(row_idx, col_name): error_message}
    """
    issues_by_cell = {}

    print(f"\nRe-validating {len(df)} rows against {len(criteria_dict)} criteria...")

    for row_idx in range(len(df)):
        row_data = df.iloc[row_idx]

        # Check each column that has criteria
        for col_name, rule in criteria_dict.items():
            if col_name not in df.columns:
                continue

            # Skip QC_FLAG column itself
            if col_name == "QC_FLAG":
                continue

            # Skip special columns - never validate them
            if col_name in SPECIAL_COLUMNS:
                continue

            value = row_data[col_name]

            # Special handling for Phase_CALC validation
            if col_name == "Phase_CALC":
                phases_value = row_data.get("PHASES", "")
                if pd.notna(phases_value):
                    phases_str = str(phases_value).strip()
                    expected_phase_calc = phase_map.get(phases_str)
                    if expected_phase_calc:
                        actual_phase_calc = str(value).strip() if pd.notna(value) else ""
                        if actual_phase_calc != expected_phase_calc:
                            issues_by_cell[(row_idx, col_name)] = f"Expected '{expected_phase_calc}' for PHASES='{phases_str}'"
                continue

            # Standard validation
            is_valid, error_msg = check_cell(value, rule, col_name, row_data)
            if not is_valid:
                issues_by_cell[(row_idx, col_name)] = error_msg

    print(f"Found {len(issues_by_cell)} cell issues after re-validation")
    return issues_by_cell


def filter_issues_by_manual_qc(issues_by_cell, non_yellow_cells):
    """
    Remove issues for cells that user has manually QC'd.

    If a cell doesn't have yellow highlighting, we consider it QC'd and won't re-flag it.
    """
    filtered_issues = {}
    manually_qcd_count = 0
    manually_qcd_examples = []

    for cell_key, error_msg in issues_by_cell.items():
        row_idx, col_name = cell_key

        # If cell is in non_yellow_cells set, user has manually QC'd it
        if cell_key in non_yellow_cells:
            manually_qcd_count += 1
            # Keep first 3 examples for debugging
            if len(manually_qcd_examples) < 3:
                manually_qcd_examples.append((row_idx + 1, col_name, error_msg))
            # Skip this issue - user has accepted/corrected it
            continue
        else:
            # Cell still has yellow or is new issue - keep it
            filtered_issues[cell_key] = error_msg

    print(f"\nManual QC detection:")
    print(f"  Total validation issues found: {len(issues_by_cell)}")
    print(f"  Cells manually QC'd (won't re-highlight): {manually_qcd_count}")
    print(f"  Cells that will remain yellow: {len(filtered_issues)}")

    # Show examples of manually QC'd cells
    if len(manually_qcd_examples) > 0:
        print(f"\n  Examples of manually QC'd cells (skipped):")
        for row_num, col_name, error_msg in manually_qcd_examples:
            print(f"    Row {row_num}, Column '{col_name}': {error_msg}")

    return filtered_issues


def add_week_number_column(df):
    """
    Add Week # column based on DATE_OUT if it doesn't exist.
    Format: YY-WXX (e.g., 24-W15)
    Uses ISO 8601 week numbering (Monday-Sunday weeks).
    Shows "N/A" for blank/invalid dates.
    """
    # Check if Week # column already exists
    if "Week #" in df.columns:
        print(f"\nWeek # column already exists, preserving existing values...")
        return

    print(f"\nAdding Week # column based on DATE_OUT...")

    week_numbers = []

    for idx, row in df.iterrows():
        date_out = row.get('DATE_OUT')

        # Check if DATE_OUT is blank or invalid
        if pd.isna(date_out):
            week_numbers.append("N/A")
        else:
            try:
                # Convert to datetime if not already
                date_obj = pd.to_datetime(date_out)

                # Get ISO week number and year
                iso_calendar = date_obj.isocalendar()
                year = iso_calendar[0]  # ISO year
                week = iso_calendar[1]  # ISO week number

                # Format as YY-WXX (e.g., 24-W15)
                year_short = year % 100  # Last 2 digits
                week_str = f"{year_short:02d}-W{week:02d}"
                week_numbers.append(week_str)

            except (ValueError, TypeError, AttributeError):
                # If conversion fails, use N/A
                week_numbers.append("N/A")

    # Insert Week # column after DATE_OUT if it exists
    if 'DATE_OUT' in df.columns:
        date_out_idx = df.columns.get_loc('DATE_OUT')
        df.insert(date_out_idx + 1, "Week #", week_numbers)
        print(f"  Added Week # column after DATE_OUT")
    else:
        df['Week #'] = week_numbers
        print(f"  Added Week # column at end")

    # Count valid weeks
    valid_weeks = sum(1 for w in week_numbers if w != "N/A")
    print(f"  Generated {valid_weeks} valid week numbers, {len(week_numbers) - valid_weeks} N/A entries")


def apply_highlighting(output_file, filtered_issues, df):
    """
    Apply yellow highlighting to cells with issues.

    IMPORTANT: This function clears ALL yellow highlighting first,
    then applies yellow ONLY to cells in filtered_issues.
    This ensures that manually un-highlighted cells stay clean.
    """
    print(f"\nApplying yellow highlighting...")

    wb = load_workbook(output_file)
    ws = wb.active

    # Get column names from header
    col_names = list(df.columns)

    # STEP 1: Clear ALL yellow highlighting first (reset)
    # This ensures cells that were manually un-highlighted stay clean
    print(f"  Step 1: Clearing all yellow highlighting...")
    cells_cleared = 0
    for row in ws.iter_rows(min_row=2, max_row=len(df)+1):
        for cell in row:
            if cell.fill and cell.fill.start_color:
                fill_color = str(cell.fill.start_color.rgb) if hasattr(cell.fill.start_color, 'rgb') else str(cell.fill.start_color.index)
                is_yellow = (
                    'FFFF00' in fill_color.upper() or
                    'FFFFFF00' in fill_color.upper() or
                    fill_color == 'FFFF0000'
                )
                if is_yellow:
                    cell.fill = NO_FILL
                    cells_cleared += 1

    print(f"    Cleared {cells_cleared} yellow cells")

    # STEP 2: Apply yellow ONLY to cells with remaining issues (filtered_issues)
    print(f"  Step 2: Applying yellow to {len(filtered_issues)} cells with issues...")
    for (row_idx, col_name), error_msg in filtered_issues.items():
        # Find column index
        if col_name in col_names:
            col_idx = col_names.index(col_name) + 1  # +1 for Excel 1-indexing
            excel_row = row_idx + 2  # +2: header row + 0-indexing

            # Get cell and apply yellow fill
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.fill = YELLOW_FILL

    # Save workbook
    wb.save(output_file)
    wb.close()

    print(f"  Highlighting applied successfully!")
    print(f"    Cells with yellow: {len(filtered_issues)}")
    print(f"    Cells cleaned: {cells_cleared - len(filtered_issues)}")


def update_qc_flag(df, filtered_issues):
    """
    Update QC_FLAG column: 1 if row has any issues, 0 if clean.
    Based ONLY on filtered_issues (which respects manual QC).
    """
    print("\nUpdating QC_FLAG column...")

    # Initialize ALL rows to QC_FLAG = 0 (clean)
    df['QC_FLAG'] = 0
    print(f"  Initialized all {len(df)} rows to QC_FLAG=0")

    # Set QC_FLAG to 1 ONLY for rows that have issues in filtered_issues
    rows_with_issues = set()
    for (row_idx, col_name) in filtered_issues.keys():
        rows_with_issues.add(row_idx)

    print(f"  Found {len(rows_with_issues)} rows with validation issues")

    # Update QC_FLAG to 1 for rows with issues
    for row_idx in rows_with_issues:
        df.at[row_idx, 'QC_FLAG'] = 1

    flagged_count = (df['QC_FLAG'] == 1).sum()
    clean_count = (df['QC_FLAG'] == 0).sum()

    print(f"  Final QC_FLAG counts:")
    print(f"    Rows with issues (QC_FLAG=1): {flagged_count}")
    print(f"    Clean rows (QC_FLAG=0): {clean_count}")

    # Verify the sum equals total rows
    if flagged_count + clean_count != len(df):
        print(f"  WARNING: QC_FLAG count mismatch! Total should be {len(df)}")

    return df


def main():
    """Main execution function."""
    print("\n" + "="*80)
    print("REVALIDATE MASTER - Script Start")
    print("="*80)

    # Find input file
    input_file = find_master_file()
    if input_file is None:
        print("\n✗ MERGE_CLEAN_QC_MASTER file not found. Exiting.")
        return

    print(f"\n  Input file: {input_file}")

    # Load QC criteria
    print("\n" + "-"*80)
    print("LOADING QC CRITERIA")
    print("-"*80)
    criteria_dict, phase_map = load_qc_criteria()
    print(f"  Loaded {len(criteria_dict)} validation rules")
    print(f"  Loaded {len(phase_map)} phase mappings")

    # Get existing highlights (cells without yellow are considered QC'd)
    non_yellow_cells = get_existing_highlights(input_file)

    # Read data
    print("\n" + "-"*80)
    print("READING DATA")
    print("-"*80)
    df = pd.read_excel(input_file)
    print(f"  Loaded {len(df)} rows, {len(df.columns)} columns")

    # Add Week # column if missing
    add_week_number_column(df)

    # Validate data
    print("\n" + "-"*80)
    print("VALIDATING DATA")
    print("-"*80)
    issues_by_cell = validate_data(df, criteria_dict, phase_map)

    # Filter out manually QC'd cells
    filtered_issues = filter_issues_by_manual_qc(issues_by_cell, non_yellow_cells)

    # Update QC_FLAG
    df = update_qc_flag(df, filtered_issues)

    # Save output file
    print("\n" + "-"*80)
    print("SAVING OUTPUT FILE")
    print("-"*80)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"MERGE_CLEAN_QC_MASTER_REVALIDATED_{timestamp}.xlsx"

    print(f"  Output file: {output_file}")

    # Convert DATE_IN and DATE_OUT to date-only format (remove time)
    print(f"  Converting date columns to date-only format...")
    date_columns = ['DATE_IN', 'DATE_OUT']
    for col in date_columns:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce').dt.date
            print(f"    Converted {col} to date-only format")

    # Save dataframe first
    df.to_excel(output_file, index=False, engine='openpyxl')

    # Apply highlighting
    if len(filtered_issues) > 0:
        apply_highlighting(output_file, filtered_issues, df)
    else:
        print("\n  No highlighting needed - all cells are clean!")

    # Summary report
    print("\n" + "="*80)
    print("SUMMARY REPORT")
    print("="*80)
    print(f"\nINPUT FILE:")
    print(f"  {input_file}")
    print(f"\nVALIDATION RESULTS:")
    print(f"  Total cells validated: {len(df) * len(criteria_dict)}")
    print(f"  Cells with issues: {len(filtered_issues)}")
    print(f"  Rows flagged (QC_FLAG=1): {(df['QC_FLAG'] == 1).sum()}")
    print(f"  Clean rows (QC_FLAG=0): {(df['QC_FLAG'] == 0).sum()}")
    print(f"\nOUTPUT FILE:")
    print(f"  {output_file}")
    print(f"\nNOTE:")
    print(f"  This MASTER_REVALIDATED file is now ready to be used as the")
    print(f"  MASTER file for the next merge cycle with Script 6.")
    print("\n" + "="*80)

    print(f"\n✓ Revalidation completed successfully!")


if __name__ == "__main__":
    main()
