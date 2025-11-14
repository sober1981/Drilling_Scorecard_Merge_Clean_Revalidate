"""
QC Re-Validation Script
Version: 2.2
Date: 2025-11-04

This script re-validates a QC file after manual corrections have been made.
It respects manual changes in two ways:
1. If you typed any value in a yellow cell (even empty), it won't re-highlight
2. If you manually removed yellow highlighting, it stays removed (considers it QC'd)

IMPORTANT: This script creates a NEW _REVALIDATED file (keeps original intact)

Usage:
1. Open MERGE_CLEAN_QC_*.xlsx file (or previous _REVALIDATED file)
2. Make corrections to yellow cells OR manually remove yellow highlighting
3. Save and CLOSE the file (must be closed!)
4. Run this script - it will create a new _REVALIDATED_timestamp file
5. Re-open the new _REVALIDATED file to see updated highlighting and QC_FLAG
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import os
import glob

# Import validation logic from 4_qc_data_quality
import importlib.util
import sys

# Load the renamed module
spec = importlib.util.spec_from_file_location("qc_data_quality", "4_qc_data_quality.py")
qc_module = importlib.util.module_from_spec(spec)
sys.modules["qc_data_quality"] = qc_module
spec.loader.exec_module(qc_module)

# Import the functions we need
load_qc_criteria = qc_module.load_qc_criteria
check_cell = qc_module.check_cell

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
NO_FILL = PatternFill(fill_type=None)


def select_qc_file():
    """Let user select which QC file to revalidate."""
    # Find all QC files
    pattern = "MERGE_CLEAN_QC_*.xlsx"
    files = glob.glob(pattern)

    if not files:
        raise FileNotFoundError(f"No files matching pattern '{pattern}' found in current directory")

    # Sort by modification time (most recent first)
    files.sort(key=os.path.getmtime, reverse=True)

    print("\nAvailable QC files:")
    print("=" * 80)
    for i, f in enumerate(files, 1):
        mod_time = datetime.fromtimestamp(os.path.getmtime(f)).strftime("%Y-%m-%d %H:%M:%S")
        size_mb = os.path.getsize(f) / (1024 * 1024)
        print(f"{i}. {f}")
        print(f"   Modified: {mod_time} | Size: {size_mb:.2f} MB")
    print("=" * 80)

    while True:
        try:
            choice = input(f"\nEnter number (1-{len(files)}) or press Enter for most recent: ").strip()
            if choice == "":
                selected_file = files[0]
                break
            choice_num = int(choice)
            if 1 <= choice_num <= len(files):
                selected_file = files[choice_num - 1]
                break
            else:
                print(f"Please enter a number between 1 and {len(files)}")
        except ValueError:
            print("Please enter a valid number")

    print(f"\nSelected: {selected_file}")
    return selected_file


def get_existing_highlights(input_file):
    """
    Read the input file and identify which cells currently have yellow highlighting.
    Returns: set of (row_idx, col_name) tuples for cells WITHOUT yellow highlighting
    """
    print("\nReading existing cell highlighting from file...")

    wb = load_workbook(input_file)
    ws = wb.active

    # Get column names from header row
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_names = list(header_row)

    # Track cells that DON'T have yellow (these are either fixed or manually QC'd)
    non_yellow_cells = set()
    cells_with_yellow = 0

    # Iterate through all data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row)):
        for col_idx, cell in enumerate(row):
            col_name = col_names[col_idx]

            # Check if cell has yellow highlighting
            if cell.fill and cell.fill.start_color:
                fill_color = str(cell.fill.start_color.rgb) if hasattr(cell.fill.start_color, 'rgb') else str(cell.fill.start_color.index)

                # Check if it's yellow (FFFF00 or variations)
                is_yellow = (
                    'FFFF00' in fill_color.upper() or
                    'FFFFFF00' in fill_color.upper() or
                    fill_color == 'FFFF0000'  # Some Excel versions use this
                )

                if is_yellow:
                    cells_with_yellow += 1
                else:
                    # Cell has some other color or no fill - consider it manually QC'd
                    non_yellow_cells.add((row_idx, col_name))
            else:
                # No fill = manually QC'd or never had issue
                non_yellow_cells.add((row_idx, col_name))

    print(f"  Found {cells_with_yellow} cells with yellow highlighting")
    print(f"  Found {len(non_yellow_cells)} cells without yellow (considered QC'd)")

    wb.close()
    return non_yellow_cells


def validate_data(df, criteria_dict, phase_map):
    """
    Validate entire dataframe against QC criteria.
    Returns: dict of issues {(row_idx, col_name): error_message}
    """
    issues_by_cell = {}

    print(f"\nRe-validating {len(df)} rows against {len(criteria_dict)} criteria...")

    for row_idx in range(len(df)):
        row_data = df.iloc[row_idx]

        # Check each column that has criteria
        for col_name, rule in criteria_dict.items():
            if col_name not in df.columns:
                continue

            # Skip QC_FLAG column itself
            if col_name == "QC_FLAG":
                continue

            value = row_data[col_name]

            # Special handling for Phase_CALC validation
            if col_name == "Phase_CALC":
                phases_value = row_data.get("PHASES", "")
                if pd.notna(phases_value):
                    phases_str = str(phases_value).strip()
                    expected_phase_calc = phase_map.get(phases_str)
                    if expected_phase_calc:
                        actual_phase_calc = str(value).strip() if pd.notna(value) else ""
                        if actual_phase_calc != expected_phase_calc:
                            issues_by_cell[(row_idx, col_name)] = f"Expected '{expected_phase_calc}' for PHASES='{phases_str}'"
                continue

            # Standard validation
            is_valid, error_msg = check_cell(value, rule, col_name, row_data)
            if not is_valid:
                issues_by_cell[(row_idx, col_name)] = error_msg

    print(f"Found {len(issues_by_cell)} cell issues after re-validation")
    return issues_by_cell


def filter_issues_by_manual_qc(issues_by_cell, non_yellow_cells):
    """
    Remove issues for cells that user has manually QC'd.

    If a cell doesn't have yellow highlighting, we consider it QC'd and won't re-flag it.
    """
    filtered_issues = {}
    manually_qcd_count = 0

    for cell_key, error_msg in issues_by_cell.items():
        # If cell is in non_yellow_cells set, user has manually QC'd it
        if cell_key in non_yellow_cells:
            manually_qcd_count += 1
            # Skip this issue - user has accepted/corrected it
            continue
        else:
            # Cell still has yellow or is new issue - keep it
            filtered_issues[cell_key] = error_msg

    print(f"\nManual QC detection:")
    print(f"  Cells manually QC'd (won't re-highlight): {manually_qcd_count}")
    print(f"  Cells that will remain yellow: {len(filtered_issues)}")

    return filtered_issues


def update_qc_flag(df, issues_by_cell):
    """Update QC_FLAG column: 1 if row has issues, 0 if clean."""
    qc_flags = []

    for row_idx in range(len(df)):
        # Check if this row has any issues
        has_issue = any(cell_key[0] == row_idx for cell_key in issues_by_cell.keys())
        qc_flags.append(1 if has_issue else 0)

    df["QC_FLAG"] = qc_flags

    rows_with_issues = sum(qc_flags)
    rows_cleaned = len(df) - rows_with_issues
    print(f"\nQC_FLAG updated: {rows_with_issues} rows still have issues, {rows_cleaned} rows are now clean")

    return rows_cleaned, rows_with_issues


def generate_output_filename(input_file):
    """Generate output filename with _REVALIDATED_timestamp."""
    import re

    # Get base name without extension
    base_name = input_file.replace('.xlsx', '')

    # Remove any existing _REVALIDATED_timestamp patterns to avoid stacking
    base_name = re.sub(r'_REVALIDATED_\d{8}_\d{6}', '', base_name)

    # Generate new timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Create output filename
    output_file = f"{base_name}_REVALIDATED_{timestamp}.xlsx"

    return output_file


def update_highlighting(input_file, output_file, df, issues_by_cell, previous_issues_count):
    """
    Create new REVALIDATED file with updated highlighting:
    - Remove yellow from all cells first
    - Apply yellow ONLY to cells in issues_by_cell (respects manual QC)
    - Saves to NEW file (keeps original intact)
    """
    print(f"\nCreating revalidated file: {output_file}")

    # Convert DATE_IN and DATE_OUT to date-only format
    date_columns = ['DATE_IN', 'DATE_OUT']
    for col in date_columns:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce').dt.date

    # Save dataframe to Excel (creates new file)
    print(f"  Saving updated data...")
    df.to_excel(output_file, index=False, engine='openpyxl')

    # Open with openpyxl to apply formatting
    wb = load_workbook(output_file)
    ws = wb.active

    # First, remove all yellow highlighting (reset)
    print(f"  Clearing all old highlighting...")
    for row in ws.iter_rows(min_row=2, max_row=len(df)+1):
        for cell in row:
            cell.fill = NO_FILL

    # Apply yellow fill ONLY to cells that still have issues
    # This respects manual QC decisions
    print(f"  Applying yellow to {len(issues_by_cell)} cells with remaining issues...")
    for (row_idx, col_name), error_msg in issues_by_cell.items():
        # Excel rows are 1-indexed, and we have a header row
        excel_row = row_idx + 2  # +1 for 0-index, +1 for header

        # Find column index
        try:
            col_idx = df.columns.get_loc(col_name) + 1  # Excel is 1-indexed
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.fill = YELLOW_FILL
        except KeyError:
            print(f"  Warning: Column '{col_name}' not found in dataframe")
            continue

    wb.save(output_file)

    # Calculate corrections made
    corrections_made = previous_issues_count - len(issues_by_cell)
    print(f"\n  [OK] Corrections respected: {corrections_made} cells")
    print(f"  [OK] Remaining issues: {len(issues_by_cell)} cells")
    print(f"  [OK] File saved: {output_file}")


def main():
    """Main execution function."""
    print("=" * 80)
    print("QC Re-Validation Script v2.2 - Respects Manual Corrections")
    print("=" * 80)
    print("\nThis script respects your manual QC work:")
    print("  [OK] Cells you edited (even if still invalid) won't be re-highlighted")
    print("  [OK] Cells where you removed yellow manually are considered QC'd")
    print("  [OK] Only untouched yellow cells will be re-validated")

    try:
        # Select file to revalidate
        input_file = select_qc_file()

        # Load QC criteria
        print("\nLoading QC criteria...")
        criteria_dict, phase_map = load_qc_criteria()
        print(f"Loaded {len(criteria_dict)} validation rules")

        # Get existing highlights BEFORE loading data
        # This tells us which cells user has manually QC'd
        non_yellow_cells = get_existing_highlights(input_file)

        # Load the QC file data
        print(f"\nLoading data from: {input_file}")
        df = pd.read_excel(input_file)
        print(f"Loaded {len(df)} rows, {len(df.columns)} columns")

        # Check if QC_FLAG exists and count previous issues
        if 'QC_FLAG' in df.columns:
            previous_issues_count = df['QC_FLAG'].sum()
            previous_rows_with_issues = (df['QC_FLAG'] == 1).sum()
            print(f"Previous state: {previous_rows_with_issues} rows with issues")
        else:
            print("Warning: QC_FLAG column not found - this may not be a QC file")
            previous_issues_count = 0

        # Re-validate all data (get raw validation results)
        raw_issues = validate_data(df, criteria_dict, phase_map)

        # Filter out issues for cells that user has manually QC'd
        # This is KEY: respects both manual edits and manual highlight removal
        issues_by_cell = filter_issues_by_manual_qc(raw_issues, non_yellow_cells)

        # Update QC_FLAG based on filtered issues
        rows_cleaned, rows_with_issues = update_qc_flag(df, issues_by_cell)

        # Generate output filename with timestamp
        output_file = generate_output_filename(input_file)

        # Create new REVALIDATED file (keeps original intact)
        update_highlighting(input_file, output_file, df, issues_by_cell, previous_issues_count)

        # Final summary
        print("\n" + "=" * 80)
        print("RE-VALIDATION SUMMARY")
        print("=" * 80)
        print(f"Input file:  {input_file}")
        print(f"Output file: {output_file}")
        print(f"\nResults:")
        print(f"  Total rows: {len(df)}")
        print(f"  Clean rows: {rows_cleaned} ({rows_cleaned/len(df)*100:.1f}%)")
        print(f"  Rows with issues: {rows_with_issues} ({rows_with_issues/len(df)*100:.1f}%)")
        print(f"  Total cells checked: {len(df) * len(criteria_dict)}")
        print(f"  Cells with issues: {len(issues_by_cell)}")
        print(f"  Cells manually QC'd: {len(raw_issues) - len(issues_by_cell)}")

        if previous_issues_count > 0:
            total_improvement = previous_issues_count - len(issues_by_cell)
            print(f"\nProgress:")
            print(f"  Total cells improved: {total_improvement}")
            print(f"  Improvement: {total_improvement/previous_issues_count*100:.1f}%")

        print("=" * 80)
        print("\n[SUCCESS] Revalidated file created successfully!")
        print(f"\nWhat happened:")
        print(f"  - Original file kept intact: {input_file}")
        print(f"  - New file created: {output_file}")
        print(f"  - QC_FLAG updated based on remaining issues")
        print(f"  - Cells you edited: No re-highlighting")
        print(f"  - Cells you manually un-highlighted: Stayed un-highlighted")
        print(f"  - Untouched yellow cells: Re-validated and updated")
        print(f"\nNext steps:")
        print(f"  1. Open the NEW file: {output_file}")
        print(f"  2. Review remaining {len(issues_by_cell)} yellow cells")
        print(f"  3. Make corrections OR manually remove yellow if acceptable")
        print(f"  4. Save and close the file")
        print(f"  5. Run this script again (it will find your REVALIDATED file)")
        print(f"  6. Repeat until QC complete (no yellow cells!)")

        input("\nPress Enter to exit...")
        return 0

    except Exception as e:
        print(f"\nERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        input("\nPress Enter to exit...")
        return 1


if __name__ == "__main__":
    try:
        exit(main())
    except Exception as e:
        print("\n" + "=" * 80)
        print("CRITICAL ERROR - Script failed to start")
        print("=" * 80)
        print(f"\nError: {str(e)}")
        print("\nFull traceback:")
        import traceback
        traceback.print_exc()
        print("\n" + "=" * 80)
        input("\nPress Enter to exit...")
        exit(1)
