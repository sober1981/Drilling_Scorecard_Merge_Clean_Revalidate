#!/usr/bin/env python3
"""
UPDATE MASTER WITH REVALIDATED - Merge Script
Version: 1.1
Date: 2025-11-24

This script updates the MASTER file with new runs from the latest REVALIDATED file.

MERGE LOGIC:
1. Duplicates are identified using THREE criteria (all must match):
   - JOB_NUM (Column D - must match exactly)
   - SN (Column AR - Serial Number must match exactly)
   - DRILLING_HOURS (Column Z - must match exactly)

2. SOURCE preservation (Column A):
   - Motor_KPI rows in MASTER: NEVER modified
   - CAM_Run_Tracker rows in MASTER: NEVER modified
   - POG_CAM_Usage rows in MASTER: NEVER modified (unless replaced by CAM_Run_Tracker)
   - POG_MM_Usage rows in MASTER: NEVER modified

3. NEW DATA from REVALIDATED:
   - New Motor_KPI rows → Added to bottom
   - New CAM_Run_Tracker rows → Added to bottom
   - New POG_CAM_Usage rows → Added to bottom (if not duplicates)
   - New POG_MM_Usage rows → Added to bottom (if not duplicates)

4. SPECIAL REPLACEMENT RULE:
   - If a CAM_Run_Tracker run in REVALIDATED matches a POG_CAM_Usage run in MASTER
     (same JOB_NUM + SN + DRILLING_HOURS), the CAM_Run_Tracker version REPLACES
     the POG_CAM_Usage version.

5. OUTPUT ORDER:
   - All original MASTER rows (except POG_CAM_Usage replaced by CAM_Run_Tracker)
   - New Motor_KPI rows from REVALIDATED
   - New CAM_Run_Tracker rows from REVALIDATED
   - New POG_CAM_Usage rows from REVALIDATED
   - New POG_MM_Usage rows from REVALIDATED

6. FORMATTING PRESERVATION:
   - Cell highlights (colors) from MASTER are preserved for MASTER rows
   - Cell highlights (colors) from REVALIDATED are preserved for new REVALIDATED rows

RESULT: Updated MASTER file with all new runs, no duplicates, proper source priority, and preserved formatting.
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import glob
import os
from copy import copy

# Configuration
SN_LAST_DIGITS = 3  # For display purposes

def find_master_file():
    """Find the MASTER file in the current directory."""
    print("\nSearching for MASTER file...")
    pattern = "*MASTER*.xlsx"
    matches = glob.glob(pattern)

    if len(matches) == 0:
        print(f"  ERROR: No file found matching pattern '{pattern}'")
        return None
    elif len(matches) > 1:
        # Sort by modification time, get most recent
        matches.sort(key=os.path.getmtime, reverse=True)
        print(f"  WARNING: Multiple files found. Using most recent: {matches[0]}")
    else:
        print(f"  Found: {matches[0]}")

    return matches[0]

def find_revalidated_file():
    """Find the most recent REVALIDATED file in the current directory."""
    print("\nSearching for REVALIDATED file...")
    pattern = "*REVALIDATED*.xlsx"
    matches = glob.glob(pattern)

    if len(matches) == 0:
        print(f"  ERROR: No file found matching pattern '{pattern}'")
        return None
    elif len(matches) > 1:
        # Sort by modification time, get most recent
        matches.sort(key=os.path.getmtime, reverse=True)
        print(f"  WARNING: Multiple files found. Using most recent: {matches[0]}")
    else:
        print(f"  Found: {matches[0]}")

    return matches[0]

def create_run_key(row):
    """
    Create a unique key for a run based on:
    - JOB_NUM (Column D)
    - SN (Column AR)
    - DRILLING_HOURS (Column Z)

    Returns:
        Tuple (job_num, sn, drilling_hours) or None if key data is missing
    """
    job_num = row['JOB_NUM']
    sn = row['SN']
    drilling_hours = row['DRILLING_HOURS']

    # Handle NaN values
    if pd.isna(job_num):
        job_num = ''
    if pd.isna(sn):
        sn = ''
    if pd.isna(drilling_hours):
        drilling_hours = 0

    # Convert to strings/numbers for consistent comparison
    job_num_str = str(job_num).strip()
    sn_str = str(sn).strip()

    # Try to convert drilling_hours to float for comparison
    try:
        drilling_hours_val = float(drilling_hours)
    except (ValueError, TypeError):
        drilling_hours_val = 0

    return (job_num_str, sn_str, drilling_hours_val)

def is_valid_run_key(run_key):
    """Check if a run key has sufficient data to be valid."""
    job_num, sn, drilling_hours = run_key

    # Need at least JOB_NUM to be valid
    if job_num == '' or job_num == 'nan':
        return False

    return True

def extract_last_digits(sn, num_digits=3):
    """Extract the last N digits from a serial number for display."""
    if pd.isna(sn):
        return ""

    sn_str = str(sn)
    digits_only = ''.join(filter(str.isdigit, sn_str))

    if len(digits_only) >= num_digits:
        return digits_only[-num_digits:]
    else:
        return digits_only

def copy_cell_format(source_cell, target_cell):
    """
    Copy formatting from source_cell to target_cell.
    Preserves: fill color, font, border, alignment, etc.
    """
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

def update_master_with_revalidated(master_file, revalidated_file):
    """
    Update MASTER file with new runs from REVALIDATED file.

    Args:
        master_file: Path to MASTER Excel file
        revalidated_file: Path to REVALIDATED Excel file

    Returns:
        Path to the output file
    """
    print("\n" + "="*80)
    print("UPDATE MASTER WITH REVALIDATED")
    print("="*80)

    # Read both files
    print("\nReading input files...")
    master_df = pd.read_excel(master_file)
    revalidated_df = pd.read_excel(revalidated_file)

    print(f"  MASTER file: {len(master_df)} rows")
    print(f"  REVALIDATED file: {len(revalidated_df)} rows")

    # Verify SOURCE column exists (Column A)
    if 'SOURCE' not in master_df.columns:
        print("\n  ERROR: SOURCE column not found in MASTER file!")
        return None
    if 'SOURCE' not in revalidated_df.columns:
        print("\n  ERROR: SOURCE column not found in REVALIDATED file!")
        return None

    # Verify required columns exist
    required_cols = ['JOB_NUM', 'SN', 'DRILLING_HOURS']
    for col in required_cols:
        if col not in master_df.columns:
            print(f"\n  ERROR: {col} column not found in MASTER file!")
            return None
        if col not in revalidated_df.columns:
            print(f"\n  ERROR: {col} column not found in REVALIDATED file!")
            return None

    print("\n  All required columns found.")

    # Display SOURCE distribution
    print("\n" + "-"*80)
    print("SOURCE DISTRIBUTION")
    print("-"*80)
    print("\nMASTER file:")
    print(master_df['SOURCE'].value_counts().to_string())
    print("\nREVALIDATED file:")
    print(revalidated_df['SOURCE'].value_counts().to_string())

    # Create run keys for all rows
    print("\n" + "-"*80)
    print("CREATING RUN KEYS")
    print("-"*80)

    master_df['RUN_KEY'] = master_df.apply(create_run_key, axis=1)
    revalidated_df['RUN_KEY'] = revalidated_df.apply(create_run_key, axis=1)

    # Build sets of run keys for duplicate detection
    master_run_keys = set()
    for idx, row in master_df.iterrows():
        run_key = row['RUN_KEY']
        if is_valid_run_key(run_key):
            master_run_keys.add(run_key)

    print(f"\n  MASTER: {len(master_run_keys)} unique valid run keys")

    revalidated_run_keys = set()
    for idx, row in revalidated_df.iterrows():
        run_key = row['RUN_KEY']
        if is_valid_run_key(run_key):
            revalidated_run_keys.add(run_key)

    print(f"  REVALIDATED: {len(revalidated_run_keys)} unique valid run keys")

    # Find overlapping run keys
    overlapping_keys = master_run_keys.intersection(revalidated_run_keys)
    print(f"  OVERLAPPING: {len(overlapping_keys)} run keys exist in both files")

    # STEP 1: Handle CAM_Run_Tracker replacing POG_CAM_Usage
    print("\n" + "-"*80)
    print("STEP 1: CAM_Run_Tracker REPLACES POG_CAM_Usage")
    print("-"*80)

    # Find CAM_Run_Tracker rows in REVALIDATED
    cam_tracker_revalidated = revalidated_df[revalidated_df['SOURCE'] == 'CAM_Run_Tracker'].copy()
    cam_tracker_keys = set()
    for idx, row in cam_tracker_revalidated.iterrows():
        run_key = row['RUN_KEY']
        if is_valid_run_key(run_key):
            cam_tracker_keys.add(run_key)

    print(f"\n  CAM_Run_Tracker in REVALIDATED: {len(cam_tracker_keys)} runs")

    # Find POG_CAM_Usage rows in MASTER that match CAM_Run_Tracker keys
    pog_cam_master = master_df[master_df['SOURCE'] == 'POG_CAM_Usage'].copy()
    replaced_pog_cam_keys = set()

    for idx, row in pog_cam_master.iterrows():
        run_key = row['RUN_KEY']
        if is_valid_run_key(run_key) and run_key in cam_tracker_keys:
            replaced_pog_cam_keys.add(run_key)

    print(f"  POG_CAM_Usage rows to be REPLACED: {len(replaced_pog_cam_keys)}")

    if len(replaced_pog_cam_keys) > 0:
        print(f"\n  REPLACEMENT DETAILS:")
        for key in replaced_pog_cam_keys:
            job_num, sn, drilling_hours = key
            sn_display = extract_last_digits(sn, SN_LAST_DIGITS)
            print(f"    JOB_NUM={job_num}, SN (last 3)={sn_display}, DRILLING_HOURS={drilling_hours:.1f}")

    # Remove POG_CAM_Usage rows from MASTER that will be replaced
    master_filtered_df = master_df[~(
        (master_df['SOURCE'] == 'POG_CAM_Usage') &
        (master_df['RUN_KEY'].apply(lambda k: is_valid_run_key(k) and k in replaced_pog_cam_keys))
    )].copy()

    print(f"\n  MASTER after POG_CAM_Usage removal: {len(master_filtered_df)} rows")

    # STEP 2: Identify NEW runs from REVALIDATED
    print("\n" + "-"*80)
    print("STEP 2: IDENTIFY NEW RUNS FROM REVALIDATED")
    print("-"*80)

    # Update master_run_keys after POG_CAM_Usage removal
    master_run_keys_updated = set()
    for idx, row in master_filtered_df.iterrows():
        run_key = row['RUN_KEY']
        if is_valid_run_key(run_key):
            master_run_keys_updated.add(run_key)

    new_runs_by_source = {
        'Motor_KPI': [],
        'CAM_Run_Tracker': [],
        'POG_CAM_Usage': [],
        'POG_MM_Usage': []
    }

    for idx, row in revalidated_df.iterrows():
        run_key = row['RUN_KEY']
        source = row['SOURCE']

        # Check if this run is new (not in updated MASTER)
        if is_valid_run_key(run_key) and run_key not in master_run_keys_updated:
            if source in new_runs_by_source:
                new_runs_by_source[source].append(idx)

    print("\n  NEW RUNS FROM REVALIDATED:")
    for source in ['Motor_KPI', 'CAM_Run_Tracker', 'POG_CAM_Usage', 'POG_MM_Usage']:
        count = len(new_runs_by_source[source])
        print(f"    {source}: {count} new runs")

    # STEP 3: Assemble the final DataFrame
    print("\n" + "-"*80)
    print("STEP 3: ASSEMBLE FINAL OUTPUT")
    print("-"*80)

    # Start with filtered MASTER (POG_CAM_Usage replaced rows removed)
    output_rows = [master_filtered_df]

    # Add new runs in order: Motor_KPI, CAM_Run_Tracker, POG_CAM_Usage, POG_MM_Usage
    for source in ['Motor_KPI', 'CAM_Run_Tracker', 'POG_CAM_Usage', 'POG_MM_Usage']:
        if len(new_runs_by_source[source]) > 0:
            new_rows = revalidated_df.loc[new_runs_by_source[source]]
            output_rows.append(new_rows)
            print(f"  Added {len(new_rows)} {source} rows")

    # Concatenate all rows
    final_df = pd.concat(output_rows, ignore_index=True)

    # Remove the temporary RUN_KEY column
    final_df = final_df.drop(columns=['RUN_KEY'])

    print(f"\n  FINAL OUTPUT: {len(final_df)} total rows")

    # Display final SOURCE distribution
    print("\n" + "-"*80)
    print("FINAL SOURCE DISTRIBUTION")
    print("-"*80)
    print(final_df['SOURCE'].value_counts().to_string())

    # Add 4 new columns at the end
    print("\n" + "-"*80)
    print("ADDING NEW COLUMNS")
    print("-"*80)

    # 1. REASON_POOH_QC - Leave blank
    final_df['REASON_POOH_QC'] = ''
    print("  Added column: REASON_POOH_QC (blank)")

    # 2. SERIES 20 - Check if SN contains "500-020"
    def check_series_20(sn):
        if pd.isna(sn):
            return 'N'
        return 'Y' if '500-020' in str(sn) else 'N'

    final_df['SERIES 20'] = final_df['SN'].apply(check_series_20)
    series_20_count = (final_df['SERIES 20'] == 'Y').sum()
    print(f"  Added column: SERIES 20 ({series_20_count} rows with 'Y', {len(final_df) - series_20_count} rows with 'N')")

    # 3. CONTROL # - Sequential unique number starting from existing max or 1
    # Check if CONTROL # already exists in master_df to continue numbering
    if 'CONTROL #' in master_df.columns:
        # Find max existing CONTROL # in master
        existing_control_nums = pd.to_numeric(master_df['CONTROL #'], errors='coerce').dropna()
        if len(existing_control_nums) > 0:
            start_num = int(existing_control_nums.max()) + 1
        else:
            start_num = 1
        print(f"  Continuing CONTROL # from existing max: starting at {start_num}")
    else:
        start_num = 1
        print(f"  Starting CONTROL # from: {start_num}")

    final_df['CONTROL #'] = range(start_num, start_num + len(final_df))
    print(f"  Added column: CONTROL # (sequential from {start_num} to {start_num + len(final_df) - 1})")

    # 4. QC BY - Leave blank
    final_df['QC BY'] = ''
    print("  Added column: QC BY (blank)")

    # Save output file WITH formatting preservation
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"MERGE_CLEAN_QC_MASTER_{timestamp}.xlsx"

    print(f"\n" + "-"*80)
    print(f"SAVING OUTPUT FILE WITH FORMATTING")
    print("-"*80)
    print(f"  Output file: {output_file}")

    # First, save the dataframe to create the Excel file structure
    final_df.to_excel(output_file, index=False, engine='openpyxl')

    # Now, open all three workbooks to copy formatting
    print(f"  Loading workbooks for formatting preservation...")
    wb_output = load_workbook(output_file)
    ws_output = wb_output.active

    wb_master = load_workbook(master_file)
    ws_master = wb_master.active

    wb_revalidated = load_workbook(revalidated_file)
    ws_revalidated = wb_revalidated.active

    # Track which rows in output came from which source
    # We need to map output rows back to source rows
    print(f"  Copying formatting from source files...")

    # Copy formatting for rows from MASTER (filtered)
    output_row = 2  # Start at row 2 (row 1 is header)

    # Rows from filtered MASTER
    for master_idx, master_row in master_filtered_df.iterrows():
        # Find this row in the original master workbook (add 2 because: 1 for header, 1 for 0-indexing)
        source_row_num = master_idx + 2

        # Copy formatting for each cell in the row
        for col_idx in range(1, len(final_df.columns) + 1):
            source_cell = ws_master.cell(row=source_row_num, column=col_idx)
            target_cell = ws_output.cell(row=output_row, column=col_idx)
            copy_cell_format(source_cell, target_cell)

        output_row += 1

    # Copy formatting for new rows from REVALIDATED (in order: Motor_KPI, CAM_Run_Tracker, POG_CAM_Usage, POG_MM_Usage)
    for source in ['Motor_KPI', 'CAM_Run_Tracker', 'POG_CAM_Usage', 'POG_MM_Usage']:
        for revalidated_idx in new_runs_by_source[source]:
            # Find this row in the original revalidated workbook
            source_row_num = revalidated_idx + 2

            # Copy formatting for each cell in the row
            for col_idx in range(1, len(final_df.columns) + 1):
                source_cell = ws_revalidated.cell(row=source_row_num, column=col_idx)
                target_cell = ws_output.cell(row=output_row, column=col_idx)
                copy_cell_format(source_cell, target_cell)

            output_row += 1

    # Save the workbook with formatting
    wb_output.save(output_file)

    # Close workbooks
    wb_master.close()
    wb_revalidated.close()
    wb_output.close()

    print(f"\n✓ Successfully saved output file with formatting preserved!")

    # Generate summary report
    print("\n" + "="*80)
    print("SUMMARY REPORT")
    print("="*80)
    print(f"\nINPUT FILES:")
    print(f"  MASTER: {master_file} ({len(master_df)} rows)")
    print(f"  REVALIDATED: {revalidated_file} ({len(revalidated_df)} rows)")
    print(f"\nOPERATIONS:")
    print(f"  POG_CAM_Usage rows REPLACED by CAM_Run_Tracker: {len(replaced_pog_cam_keys)}")
    print(f"  Motor_KPI rows ADDED: {len(new_runs_by_source['Motor_KPI'])}")
    print(f"  CAM_Run_Tracker rows ADDED: {len(new_runs_by_source['CAM_Run_Tracker'])}")
    print(f"  POG_CAM_Usage rows ADDED: {len(new_runs_by_source['POG_CAM_Usage'])}")
    print(f"  POG_MM_Usage rows ADDED: {len(new_runs_by_source['POG_MM_Usage'])}")
    print(f"\nOUTPUT FILE:")
    print(f"  {output_file} ({len(final_df)} rows)")
    print(f"  Net change: {len(final_df) - len(master_df):+d} rows")
    print("\n" + "="*80)

    return output_file

def main():
    """Main execution function."""
    print("\n" + "="*80)
    print("UPDATE MASTER WITH REVALIDATED - Script Start")
    print("="*80)

    # Find input files
    master_file = find_master_file()
    if master_file is None:
        print("\n✗ MASTER file not found. Exiting.")
        return

    revalidated_file = find_revalidated_file()
    if revalidated_file is None:
        print("\n✗ REVALIDATED file not found. Exiting.")
        return

    # Update master with revalidated
    output_file = update_master_with_revalidated(master_file, revalidated_file)

    if output_file:
        print(f"\n✓ Process completed successfully!")
        print(f"✓ Output file: {output_file}")
    else:
        print("\n✗ Process failed.")

if __name__ == "__main__":
    main()
